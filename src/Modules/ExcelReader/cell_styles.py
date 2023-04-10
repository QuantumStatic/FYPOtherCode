from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection, NamedStyle, Color
from other_utils import object_merger, make_list_linear

def Center_align():
    center_align  = NamedStyle(name="Center Align")
    center_align.alignment = Alignment(horizontal='center', vertical='center')
    return center_align

def Wrap_text():
    wrap_text  = NamedStyle(name="Wrap Text")
    wrap_text.alignment = Alignment(wrap_text=True)
    return wrap_text

def Bold():
    bold  = NamedStyle(name="Bold")
    bold.font = Font(bold=True)
    return bold

def Italic():
    italic  = NamedStyle(name="Italic")
    italic.font = Font(italic=True)
    return italic

def Fully_bordered():
    fully_bordered = NamedStyle(name="Fully Bordered")
    border_style = "thin"
    fully_bordered.borders = Border(left=Side(style=border_style), right=Side(style=border_style), top=Side(style=border_style), bottom=Side(style=border_style))
    return fully_bordered

def No_border():
    no_border = NamedStyle(name="No Border")
    no_border.borders = Border(left=Side(style='none'), right=Side(style='none'), top=Side(style='none'), bottom=Side(style='none'))
    return no_border

def Font_colour_amber():
    amber = NamedStyle(name="Font Colour Amber")
    amber.font = Font(color='FCC000')
    return amber

def Font_colour_green():
    green = NamedStyle(name="Font Colour Green")
    green.font = Font(color='00B05  0')
    return green

def Font_colour_red():
    red = NamedStyle(name="Font Colour Red")
    red.font = Font(color='FF0000')
    return red

def Font_colour_blue():
    blue = NamedStyle(name="Font Colour Blue")
    blue.font = Font(color='0070C0')
    return blue

def Font_colour_black():
    black = NamedStyle(name="Font Colour Black")
    black.font = Font(color='000000')
    return black

def Font_colour_white():
    white = NamedStyle(name="Font Colour White")
    white.font = Font(color='FFFFFF')
    return white

def Cell_colour_light_blue():
    colour = Color("4287f5")
    cell_colour_light_blue = NamedStyle(name="Cell Colour Light Blue")
    cell_colour_light_blue.fill = PatternFill(fill_type='solid', start_color=colour, end_color=colour, patternType='solid')
    return cell_colour_light_blue

def Cell_colour_dark_blue():
    colour = Color("0070C0")
    cell_colour_dark_blue = NamedStyle(name="Cell Colour Dark Blue")
    cell_colour_dark_blue.fill = PatternFill(fill_type='solid', start_color=colour, end_color=colour, patternType='solid')
    return cell_colour_dark_blue

def Cell_colour_green():
    colour = Color("70AD47")
    cell_colour_green = NamedStyle(name="Cell Colour Green")
    cell_colour_green.fill = PatternFill(fill_type='solid', start_color=colour, end_color=colour, patternType='solid')
    return cell_colour_green

def Cell_colour_red():
    colour = Color("FF0000")
    Cell_colour_red = NamedStyle(name="Cell Colour Red")
    Cell_colour_red.fill = PatternFill(fill_type='solid', start_color=colour, end_color=colour, patternType='solid')
    return Cell_colour_red

def Coloumn_header():
    coloumn_header = NamedStyle(name="Coloumn Header")
    coloumn_header.font = Font_colour_white().font
    coloumn_header.bold = True
    coloumn_header.alignment = Center_align().alignment
    coloumn_header.alignment.wrap_text = True
    coloumn_header.border = Fully_bordered().border
    return coloumn_header

def get_style(name:str):
    name = name.replace('-', '_')
    name = name.replace(' ', '_')
    name = name.capitalize()
    return eval(name)()

def _merge_styles(*to_merge) -> NamedStyle:
    to_merge = make_list_linear(to_merge)
    styles_to_merge = [get_style(style) for style in to_merge]
    return object_merger(styles_to_merge)
