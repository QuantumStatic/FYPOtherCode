from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Any,  Union, NamedTuple
import re
from collections import namedtuple
from ExcelAutoExceptions import PermissionDenied
from openpyxl.styles import NamedStyle, Font, Alignment
from openpyxl.styles.colors import Color
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.protection import Protection


Cell: NamedTuple = NamedTuple('Cell', [('row', int), ('col', Union[int, str])])

def _col_conversion(col: Union[str, int]) -> Union[str, int]:
    if isinstance(col, int):
        return get_column_letter(col)
    elif isinstance(col, str):
        if col.isnumeric():
            return int(col)
        return column_index_from_string(col)
    else:
        raise TypeError(f"{type(col)} is not supported. col must be either str or int")

def _extract_col(cell: str, default: Any = '-1') -> str:
    if cell.find(' ') == -1:
        try:
            row_begin = re.compile(r'[A-Z]+(\d)').findall(cell)[0]
        except IndexError:
            return cell if not cell.isnumeric() else default
        else:
            return cell[:cell.find(row_begin)]
    else:
        return cell.split()[0]

def _extract_row(cell: str, default: Any = '-1') -> str:
    if cell.find(' ') == -1:
        try:
            row_begin = re.compile(r'[A-Z]+(\d)').findall(cell)[0]
        except IndexError:
            return cell if cell.isnumeric() else default
        else:
            return cell[cell.find(row_begin):]
    else:
        return cell.split()[1]

def _cellStr_to_Cell(cell: str) -> Cell:
    return Cell(row=int(_extract_row(cell)), col=_extract_col(cell))

def convert_cell(cell_str: str = None, row: int = '', col: Union[int, str] = '', return_cell_str: bool = True, only_col: bool = None, only_row: bool = None, return_type: type = None) -> Union[str, Cell]:
    if cell_str is None:
        if row == '' and col is not None:
            only_col = True
            only_row = False
        elif col == '' and row is not None:
            only_col = False
            only_row = True
        cell_str  = f"{col} {row}".strip()
    else:
        if isinstance(cell_str, tuple) or isinstance(cell_str, list):
            cell_str = f"{cell_str[0]} {cell_str[1]}".strip()
        if (col := _extract_col(cell_str, False)) is not False and len(col) == len(cell_str):
            only_col = True
            only_row = None
        elif (row := _extract_row(cell_str, False)) is not False and len(row) == len(cell_str):
            only_col = None
            only_row = True
        
        assert not only_col or not only_row, "only_col and only_row cannot be both True. Set only 1 of them to True"

        if only_col is None and only_row is None:        
            cell = _cellStr_to_Cell(cell_str)

            if not isinstance(cell.col, str):
                cell = Cell(col=_col_conversion(cell.col), row=cell.row)
            if cell.col.isnumeric():
                cell = Cell(col=_col_conversion(int(cell.col)), row=cell.row)
            
            if return_cell_str:
                return f"{cell.col}{cell.row}"
            else:
                return Cell(col=_col_conversion(cell.col), row=cell.row)
        
        elif only_col:
            Col = _extract_col(cell_str, False)

            if Col is False and return_type is str:
                return _col_conversion(int(cell_str))
            elif isinstance(col, str) and return_type is int:
                return _col_conversion(Col)
            else:
                return int(cell_str) if cell_str.isnumeric() else Col
        
        elif only_row:
            Row = _extract_row(cell_str, False)
            return return_type(Row)

def make_list_linear(list_to_squish: Union[list, tuple]) -> list:
    final_list = []
    for item in list_to_squish:
        if isinstance(item, list) or isinstance(item, tuple):
            final_list.extend(make_list_linear(item))
        else:
            final_list.append(item)
    return final_list

def object_merger(*to_merge):
    attributes = make_list_linear(to_merge)
    combined_object = eval(type(attributes[0]).__name__)()
    total_sz = len(attributes)

    if total_sz == 1:
        combined_object = attributes[0]
    
    while total_sz > 1:
        for attrs in attributes[0].__dict__:
            if hasattr(attributes[0].__dict__[attrs], '__dict__') and hasattr(attributes[1].__dict__[attrs], '__dict__'):
                attrs_to_pass = (attributes[0].__dict__[attrs], attributes[1].__dict__[attrs])
                combined_object.__dict__[attrs] = object_merger(*attrs_to_pass)
            else:
                if ((curr_val := attributes[0].__dict__[attrs]) is not None and curr_val  is not False and not (isinstance(curr_val, str) and curr_val.isnumeric())) or (isinstance(curr_val, str) and curr_val.isnumeric() and int(curr_val) >0):
                    combined_object.__dict__[attrs] = curr_val
                elif (curr_val := attributes[1].__dict__[attrs]) is not None:
                    combined_object.__dict__[attrs] = curr_val
        attributes = attributes[1:]
        attributes[0] = combined_object
        total_sz -= 1
    return combined_object