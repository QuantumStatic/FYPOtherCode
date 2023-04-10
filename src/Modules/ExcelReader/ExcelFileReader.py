from __future__ import annotations
import styling
import cell_styles
from other_utils import convert_cell, make_list_linear

from openpyxl.styles import NamedStyle
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.copier import WorksheetCopy
from openpyxl import load_workbook, worksheet, Workbook
from ExcelAutoExceptions import SaveFailedException, WorksheetNotFoundException, WorkbookCloseFailedException

from zipfile import BadZipFile
import gc
from typing import Union, Any
from functools import reduce
import os


class ExcelFileReader:
    open_files: list[ExcelFileReader] = []

    def __init__(self, path: str, using_manager: bool = True, data_only: bool = False):
        self._path = path
        self._col_headers = {}
        self._styles: set[str] = set()
        self._save_tries = 0
        self._data_only = data_only
        self._ws_deets: dict[str, dict] = {}

        ExcelFileReader.open_files.append(self)

        if not using_manager:
            self._setup()

    def _setup(self):
        try:
            self._workbook = load_workbook(self._path, data_only=self._data_only)
        except FileNotFoundError:
            print(f"No file named {self.name_of_file} was found. Creating a new file at {self._path}.")
            self._workbook = Workbook()
        except BadZipFile:
            print(f"{type(self).__name__} does not support locked files.")
        else:
            print("successfully loaded workbook")

        self._current_worksheet: worksheet = self._workbook.active
        self._ws_deets[self._current_worksheet.title] = {}

    def __enter__(self):
        self._setup()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._workbook.close()

    def __call__(self):
        return NotImplemented

    def __getitem__(self, cell: Union[str, tuple[int, int], tuple[str, int]]) -> Cell:
        try:
            if isinstance(cell, tuple) and isinstance(cell[0], int):
                return self._current_worksheet.cell(column=cell[0], row=cell[1])

            return self._current_worksheet[convert_cell(cell_str=cell)]

        except ValueError as error:
            print(error)

    def __setitem__(self, cell: Union[str, tuple[int, int], tuple[str, int]], value: Any):
        try:
            if isinstance(cell, tuple) and isinstance(cell[0], int):
                self._current_worksheet.cell(
                    column=cell[0], row=cell[1]).value = value

            self._current_worksheet[convert_cell(cell_str=cell)].value = value

        except ValueError as error:
            print(error)

    def change_active_worksheet(self, new_worksheet_name: str, mode: str = None, no_create: bool = True):
        try:
            self._current_worksheet = self._workbook[new_worksheet_name]
        except KeyError as error:
            if no_create:
                raise WorksheetNotFoundException(new_worksheet_name) from error
            self._current_worksheet = self._workbook.create_sheet(
                new_worksheet_name)
        finally:
            self._ws_deets[new_worksheet_name] = {}

        return True

    def save(self, path: str = None, relative=False):
        if path is not None:
            new_path = ('\\'.join(self._path.split('\\')[
                        :-1]) + '\\' + path) if relative else path
        else:
            new_path = self._path

        try:
            try:
                os.remove(new_path)
            except FileNotFoundError:
                pass
            finally:
                self._workbook.save(new_path)
        except PermissionError:
            print("It appears that the file you are tying to modify is still open. Please close the file and try again.")
            input("After closing the file, press any key to continue.")
            self._save_tries += 1
            if self._save_tries > 5:
                raise SaveFailedException
            self.save(path, relative)
        else:
            self._save_tries = 0
            print("Successfully saved file.")

    def _close(self):
        try:
            self._workbook.close()
            gc.collect()
        except Exception as error:
            raise WorkbookCloseFailedException(error) from error

    def column_headers(self, offset: int = 0):
        self._col_headers = {}
        if offset == 0 and self._ws_deets[self._current_worksheet.title].setdefault('header_offset', False) is not False:
            offset = self._ws_deets[self._current_worksheet.title]['header_offset']
        else:
            self._ws_deets[self._current_worksheet.title]['header_offset'] = offset

        for col in range(self._current_worksheet.max_column):
            self._col_headers[self._current_worksheet.cell(
                column=col+1, row=offset+1).value] = col

        return self._col_headers

    @property
    def rows(self):
        rows = self._current_worksheet.rows
        for _ in range(self._ws_deets[self._current_worksheet.title].setdefault('header_offset', 0)+1):
            next(rows)
        return rows

    @property
    def worksheets(self):
        return self._workbook.sheetnames

    def copy_current_worksheet(self, new_worksheet_name: str, no_create: bool = False):
        self.copy_worksheet(self._current_worksheet.title,
                            new_worksheet_name, no_create)

    def copy_worksheet(self, worksheet_to_copy: str, new_worksheet: str, no_create: bool = False):
        if no_create:
            if new_worksheet not in self._workbook.sheetnames:
                raise WorksheetNotFoundException(new_worksheet)
            else:
                new_worksheet_name = new_worksheet
        else:
            new_worksheet_name = new_worksheet if new_worksheet not in self._workbook.sheetnames else new_worksheet + '_copy'

        copy_ws = WorksheetCopy(
            self._workbook[worksheet_to_copy], self._workbook[new_worksheet_name])
        copy_ws.copy_worksheet()

        return True

    def _register_cell_style(self, style: NamedStyle):
        try:
            self._styles.add(style.name)
            self._workbook.add_named_style(style)
        except ValueError:
            pass

    def add_style(self, *to_merge):
        to_merge = make_list_linear(to_merge)
        if len(to_merge) > 1:
            name_str = reduce(
                lambda x, y: f"{x} {y} ", sorted(to_merge)).strip()
        else:
            name_str = to_merge[0]

        if name_str not in self._styles:
            new_style = cell_styles._merge_styles(to_merge)
            new_style.name = name_str
            self._register_cell_style(new_style)

        return name_str

    def style_adjacent_cells(self, start_cell: str, end_cell: str, *style_string: str):
        styling.style_adjacent_cells(self, start_cell, end_cell, style_string)

    def style_cols(self, col_str: Union[str, int, list[int, str], tuple[int, str]], *style_string: str, skip_init_rows: int = 0):
        styling.style_cols(self, col_str, style_string,
                           skip_init_rows=skip_init_rows)

    def style_rows(self, row_str: Union[str, int, list[int, str], tuple[int, str]], *style_string: str, skip_init_cols: int = 0):
        styling.style_rows(self, row_str, style_string,
                           skip_init_rows=skip_init_cols)

    @property
    def total_columns(self):
        return self._current_worksheet.max_column

    @property
    def total_rows(self):
        return self._current_worksheet.max_row

    @property
    def worksheet_name(self):
        return self._current_worksheet.title

    def insert_col(self, col_index: Union[str, int], col_name: str = None):
        self._current_worksheet.insert_cols(
            convert_cell(col=col_index, return_type=int), amount=1)
        self._current_worksheet[col_index + '1'].value = col_name

    def delete_row(self, row: int):
        self._current_worksheet.delete_rows(row)

    def delete_rows(self, start: int, end: int):
        rows_to_del = end - start
        self._current_worksheet.delete_rows(start, rows_to_del)

    def delete_col(self, col: Union[str, int]):
        self._current_worksheet.delete_cols(
            convert_cell(col=col, return_type=int), amount=1)

    def delete_cols(self, start: Union[str, int], stop: Union[str, int]):
        start = convert_cell(col=start, return_type=int)
        rows_to_del = convert_cell(col=stop, return_type=int) - start
        self._current_worksheet.delete_cols(start, rows_to_del)

    def unmerge_cells(self, start: str, end: str):
        self._current_worksheet.unmerge_cells(f"{start}:{end}")

    def unmerge_all(self):
        while len(self._current_worksheet.merged_cells.ranges) > 0:
            for merged_cells in self._current_worksheet.merged_cells.ranges:
                self._current_worksheet.unmerge_cells(str(merged_cells))

    def _full_col(self, column: str):
        return f"{column}1:{column}{self._current_worksheet.max_row}"

    def _full_row(self, row: int):
        return f"A{row}:{self._current_worksheet.max_column}{row}"

    def shift_row(self, start_row: int, amount: int, end_row: int = None):
        if end_row is not None:
            total_rows_shift_str = f"{self._full_row(start_row).split(':')[0]}:{self._full_col(end_row).split(':')[1]}"
            total_rows_shift_int = end_row - start_row
        else:
            total_rows_shift_str = self._full_row(start_row)
            total_rows_shift_int = 1

        distance_from_last_row = self._current_worksheet.max_row - start_row
        self._current_worksheet.move_range(
            cell_range=total_rows_shift_str, rows=distance_from_last_row+1)

        if amount > 0:
            other_rows_shift_beg = f"{end_row + 1}"
            other_rows_shift_end = f"{end_row + amount}"
            interim_shift = -1 * total_rows_shift_int
        else:
            other_rows_shift_beg = f"{start_row + amount}"
            other_rows_shift_end = f"{start_row - 1}"
            interim_shift = 1 * total_rows_shift_int

        other_rows_shift = f"A:{other_rows_shift_beg}:{self._current_worksheet.max_column}{other_rows_shift_end}"
        self._current_worksheet.move_range(
            cell_range=other_rows_shift, rows=interim_shift)

        new_total_rows_shift_str = self._full_row(
            start_row + distance_from_last_row + 1)
        if end_row is not None:
            new_total_rows_shift_str = f"{self._full_row(start_row + distance_from_last_row + 1).split(':')[0]}:{self._full_row(start_row + distance_from_last_row + total_rows_shift_int).split(':')[1]}"

        self._current_worksheet.move_range(
            cell_range=new_total_rows_shift_str, rows=-1 * distance_from_last_row + amount - 1)

    def shift_column(self, start_col: str, amount: int, end_col: str = None):
        start_col_int, start_col_str = convert_cell(
            col=start_col, return_type=int), convert_cell(col=start_col, return_type=str)

        total_cols_shift_str, total_cols_shift_int = self._full_col(
            start_col_str), 1
        end_col_int = start_col_int

        if end_col is not None:
            end_col_int, end_col_str = convert_cell(
                col=end_col, return_type=int), convert_cell(col=end_col, return_type=str)

            total_cols_shift_str, total_cols_shift_int = f"{total_cols_shift_str.split(':')[0]}:{self._full_col(end_col_str).split(':'[1])}", end_col_int - \
                start_col_int + 1

        distance_from_last_col = self._current_worksheet.max_column - end_col_int
        self._current_worksheet.move_range(
            cell_range=total_cols_shift_str, columns=distance_from_last_col+amount)

        # Special case for the last column
        if distance_from_last_col == 0 and amount > 0:
            return

        if amount > 0:
            other_cols_shift_beg = convert_cell(
                col=end_col_int + 1, return_type=str)
            other_cols_shift_end = convert_cell(
                col=end_col_int + amount, return_type=str)
            interim_shift = -1 * total_cols_shift_int
        else:
            other_cols_shift_beg = convert_cell(
                col=start_col_int + amount, return_type=str)
            other_cols_shift_end = convert_cell(
                col=start_col_int - 1, return_type=str)
            interim_shift = 1 * total_cols_shift_int

        if amount < 0 or end_col_int < self.total_columns:
            other_cols_shift = f"{other_cols_shift_beg}1:{other_cols_shift_end}{self._current_worksheet.max_row}"
            self._current_worksheet.move_range(
                cell_range=other_cols_shift, columns=interim_shift)

        new_total_cols_shift_str = f"{self._full_col(convert_cell(col=start_col_int + distance_from_last_col + 1, return_type=str))}"
        if end_col is not None:
            new_start_col_str, new_end_col_str = convert_cell(col=start_col_int + distance_from_last_col + 1, return_type=str), convert_cell(
                col=start_col_int + distance_from_last_col + total_cols_shift_int, return_type=str)
            new_total_cols_shift_str = f"{new_start_col_str}:{new_end_col_str}"

        self._current_worksheet.move_range(
            cell_range=new_total_cols_shift_str, columns=-1 * distance_from_last_col + amount - 1)

    def shift_all_cells_rowise(self, amount: int = 1):
        self._current_worksheet.move_range(
            cell_range=f"A1:{convert_cell(col=self.total_columns, return_type=str)}{self.total_rows}", rows=amount)

    def swap_col(self, col1: Union[int, str], col2: Union[int, str]):
        col1 = convert_cell(col=col1, return_type=str)
        col2 = convert_cell(col=col2, return_type=str)
        col1_loc_int, col2_loc_int = convert_cell(col=col1, only_col=True, return_type=int), convert_cell(
            col=col2, only_col=True, return_type=int)
        col1 = self._full_col(col1)
        col2 = self._full_col(col2)

        self._current_worksheet.move_range(
            cell_range=col1, columns=self.total_columns - col1_loc_int+1)

        col1 = convert_cell(col=self.total_columns,
                            only_col=True, return_type=str)
        col1 = self._full_col(col1)

        self._current_worksheet.move_range(
            cell_range=col2, cols=col1_loc_int-col2_loc_int)
        self._current_worksheet.move_range(
            cell_range=col1, cols=col2_loc_int-self.total_columns)

    @property
    def name_of_file(self) -> str:
        return self._path.split('\\')[-1].split('.')[0]
